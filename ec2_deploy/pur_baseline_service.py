"""
Pur Baseline Service
====================
Handles:
  1. Upload files to S3 (multiple prefixes)
  2. Parse files and store results in EC2 SQLite (tutti_pur_baseline.db)

Tables:
  - tuttiprequn: Marker pre-equation (from Tutti_各marker給線Template)
  - tuttiprueqn: Production equation (from 生產給線)
  - tuttirealassign: Real assignment data (from IDEXX 外測值總表*.xlsx)

Equation CSV Format (each file = 1 marker):
    Marker Name,BA
    Well Number,5
    Control Equation,y = 391.303x  -1.459
    Canine Equation,y = 391.303x  -1.459
    Feline Equation,y = 391.303x  -1.459
    Equine Equation,y = 391.303x  -1.459
"""

import csv
import io
import logging
import os
import re
import sqlite3
from datetime import datetime, timezone
from pathlib import Path

import boto3
from botocore.exceptions import ClientError

logger = logging.getLogger(__name__)

# ─── Config ───────────────────────────────────────────────────────────────────

S3_BUCKET = os.getenv("PUR_BASELINE_S3_BUCKET", "beads-photos-harry")
S3_REGION = os.getenv("PUR_BASELINE_S3_REGION", "ap-northeast-1")

# SQLite DB path on EC2
PUR_BASELINE_DB_PATH = Path(
    os.getenv(
        "PUR_BASELINE_DB_PATH",
        "/home/ubuntu/qc-web-ipqc/tutti-qc-assayprocess/data/tutti_pur_baseline.db",
    )
)


# ─── S3 Upload ────────────────────────────────────────────────────────────────

_s3_client = None


def get_s3_client():
    global _s3_client
    if _s3_client is None:
        _s3_client = boto3.client("s3", region_name=S3_REGION)
    return _s3_client


def upload_file_to_s3(
    file_content: bytes,
    s3_bucket: str,
    s3_key: str,
) -> dict:
    """Upload bytes to S3. Returns result dict."""
    try:
        client = get_s3_client()
        client.put_object(
            Bucket=s3_bucket,
            Key=s3_key,
            Body=file_content,
        )
        s3_url = f"https://{s3_bucket}.s3.{S3_REGION}.amazonaws.com/{s3_key}"
        logger.info(f"S3 upload OK: s3://{s3_bucket}/{s3_key}")
        return {"ok": True, "s3_key": s3_key, "s3_url": s3_url}
    except ClientError as exc:
        error_msg = str(exc)
        logger.error(f"S3 upload failed: {error_msg}")
        return {"ok": False, "error": error_msg}


# ─── SQLite DB ────────────────────────────────────────────────────────────────

SCHEMA_SQL = """
-- Table 1: marker pre-equation (各marker給線Template)
CREATE TABLE IF NOT EXISTS tuttiprequn (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    marker_name TEXT NOT NULL,
    well_number TEXT,
    control_equation TEXT,
    canine_equation TEXT,
    feline_equation TEXT,
    equine_equation TEXT,
    source_file TEXT,
    file_hash TEXT,
    s3_key TEXT,
    s3_url TEXT,
    created_at TEXT NOT NULL DEFAULT (datetime('now')),
    updated_at TEXT NOT NULL DEFAULT (datetime('now')),
    UNIQUE(marker_name, source_file)
);

-- Table 2: production equation (生產給線)
CREATE TABLE IF NOT EXISTS tuttiprueqn (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    marker_name TEXT NOT NULL,
    well_number TEXT,
    control_equation TEXT,
    canine_equation TEXT,
    feline_equation TEXT,
    equine_equation TEXT,
    production_date TEXT,
    disc_name TEXT,
    sub_panel_type TEXT,
    source_file TEXT,
    file_hash TEXT,
    s3_key TEXT,
    s3_url TEXT,
    created_at TEXT NOT NULL DEFAULT (datetime('now')),
    updated_at TEXT NOT NULL DEFAULT (datetime('now')),
    UNIQUE(marker_name, source_file)
);

-- Table 3: real assignment (IDEXX 外測值總表)
CREATE TABLE IF NOT EXISTS tuttirealassign (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    marker_name TEXT,
    species TEXT,
    sample_id TEXT,
    expected_value TEXT,
    measured_value TEXT,
    unit TEXT,
    result_status TEXT,
    sheet_name TEXT,
    row_index INTEGER,
    source_file TEXT,
    file_hash TEXT,
    s3_key TEXT,
    s3_url TEXT,
    raw_data TEXT,
    created_at TEXT NOT NULL DEFAULT (datetime('now')),
    updated_at TEXT NOT NULL DEFAULT (datetime('now'))
);

-- Table 4: FW 計算值 (Tutti-計算值.xlsx)
CREATE TABLE IF NOT EXISTS calvalue (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    marker_name TEXT,
    parameter TEXT,
    value TEXT,
    unit TEXT,
    sheet_name TEXT,
    row_index INTEGER,
    source_file TEXT,
    file_hash TEXT,
    s3_key TEXT,
    s3_url TEXT,
    raw_data TEXT,
    created_at TEXT NOT NULL DEFAULT (datetime('now')),
    updated_at TEXT NOT NULL DEFAULT (datetime('now'))
);

CREATE INDEX IF NOT EXISTS idx_tuttiprequn_marker ON tuttiprequn(marker_name);
CREATE INDEX IF NOT EXISTS idx_tuttiprueqn_marker ON tuttiprueqn(marker_name);
CREATE INDEX IF NOT EXISTS idx_tuttirealassign_marker ON tuttirealassign(marker_name);
CREATE INDEX IF NOT EXISTS idx_tuttirealassign_source ON tuttirealassign(source_file);
CREATE INDEX IF NOT EXISTS idx_calvalue_marker ON calvalue(marker_name);
CREATE INDEX IF NOT EXISTS idx_calvalue_source ON calvalue(source_file);
"""


def _get_db() -> sqlite3.Connection:
    PUR_BASELINE_DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(str(PUR_BASELINE_DB_PATH))
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA busy_timeout=5000")
    return conn


def init_pur_baseline_db() -> None:
    conn = _get_db()
    conn.executescript(SCHEMA_SQL)
    # Migrate: add columns if they don't exist (for existing DBs)
    try:
        conn.execute("ALTER TABLE tuttiprueqn ADD COLUMN production_date TEXT")
    except sqlite3.OperationalError:
        pass
    try:
        conn.execute("ALTER TABLE tuttiprueqn ADD COLUMN disc_name TEXT")
    except sqlite3.OperationalError:
        pass
    try:
        conn.execute("ALTER TABLE tuttiprueqn ADD COLUMN sub_panel_type TEXT")
    except sqlite3.OperationalError:
        pass
    conn.close()
    logger.info(f"Pur baseline DB initialized: {PUR_BASELINE_DB_PATH}")


# ─── Path Parsing (生產給線) ──────────────────────────────────────────────────

# Pattern: .../YYYYMMDD_DISCNAME 給線For 生產/...
_PRODUCTION_DIR_PATTERN = re.compile(r"(\d{8})_(\S+)\s*給線")


def extract_production_info(source_path: str) -> tuple[str, str]:
    """
    Extract production_date and disc_name from source file path.
    Example: /mnt/reagent_rd/0. Tutti/生產給線/2026/20260429_OW3 給線For 生產/Equation...
    Returns: ("20260429", "OW3") or ("", "")
    """
    match = _PRODUCTION_DIR_PATTERN.search(source_path)
    if match:
        return match.group(1), match.group(2)
    return "", ""


# ─── RDS Lookup (sub_panel_type) ──────────────────────────────────────────────

_disc_to_subpanel_cache: dict[str, str] | None = None


def _load_disc_to_subpanel() -> dict[str, str]:
    """Load disc_name → sub_panel_type mapping from RDS qbi_qr.paneltype."""
    global _disc_to_subpanel_cache
    if _disc_to_subpanel_cache is not None:
        return _disc_to_subpanel_cache

    try:
        import psycopg2
        from dotenv import load_dotenv
        load_dotenv()
        rds_config = {
            "host": os.getenv("TUTTI_RDS_HOST", "database-1.cfutwrwyrxts.ap-northeast-1.rds.amazonaws.com"),
            "port": int(os.getenv("TUTTI_RDS_PORT", "5432")),
            "database": os.getenv("TUTTI_RDS_DATABASE", "beadsdb"),
            "user": os.getenv("TUTTI_RDS_USER", "harryguo"),
            "password": os.getenv("TUTTI_RDS_PASSWORD", "skyla168"),
        }
        pg = psycopg2.connect(**rds_config)
        cur = pg.cursor()
        cur.execute("SELECT disc_name, sub_panel_type FROM qbi_qr.paneltype WHERE disc_name IS NOT NULL AND disc_name != ''")
        mapping = {}
        for row in cur.fetchall():
            disc = row[0].strip()
            sub = row[1].strip() if row[1] else ""
            if disc:
                mapping[disc] = sub
                # Also store uppercase version for case-insensitive lookup
                mapping[disc.upper()] = sub
        cur.close()
        pg.close()
        _disc_to_subpanel_cache = mapping
        logger.info(f"Loaded {len(mapping)//2} disc_name → sub_panel_type mappings from RDS")
        return mapping
    except Exception as exc:
        logger.warning(f"Failed to load disc→subpanel from RDS: {exc}")
        _disc_to_subpanel_cache = {}
        return {}


def lookup_sub_panel_type(disc_name: str) -> str:
    """Look up sub_panel_type from disc_name via RDS qbi_qr.paneltype."""
    if not disc_name:
        return ""
    mapping = _load_disc_to_subpanel()
    # Try exact match first, then uppercase
    return mapping.get(disc_name, mapping.get(disc_name.upper(), ""))


# ─── CSV Parsing (Equation files) ────────────────────────────────────────────

def parse_equation_csv(content: bytes) -> dict | None:
    """
    Parse a single Equation CSV file (key,value per row).
    Returns dict with marker data, or None if parsing fails.
    """
    try:
        try:
            text = content.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = content.decode("cp950")

        reader = csv.reader(io.StringIO(text))
        data = {}
        for row in reader:
            if len(row) >= 2:
                key = row[0].strip()
                value = ",".join(row[1:]).strip()  # Handle commas in equation
                data[key] = value

        marker_name = data.get("Marker Name", "").strip()
        if not marker_name:
            return None

        return {
            "marker_name": marker_name,
            "well_number": data.get("Well Number", ""),
            "control_equation": data.get("Control Equation", ""),
            "canine_equation": data.get("Canine Equation", ""),
            "feline_equation": data.get("Feline Equation", ""),
            "equine_equation": data.get("Equine Equation", ""),
        }
    except Exception as exc:
        logger.warning(f"Failed to parse CSV: {exc}")
        return None


# ─── XLSX Parsing (外測值總表) ────────────────────────────────────────────────

def parse_real_assign_xlsx(content: bytes, file_name: str) -> list[dict]:
    """
    Parse IDEXX 外測值總表 xlsx file.
    Returns list of row dicts for tuttirealassign table.
    """
    try:
        import openpyxl
        import io as _io

        wb = openpyxl.load_workbook(_io.BytesIO(content), read_only=True, data_only=True)
        rows = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            all_rows = list(ws.iter_rows(values_only=True))
            if len(all_rows) < 2:
                continue

            # First row is header
            headers = [str(h or "").strip() for h in all_rows[0]]

            for row_idx, row_values in enumerate(all_rows[1:], start=2):
                row_dict = {}
                for col_idx, value in enumerate(row_values):
                    if col_idx < len(headers) and headers[col_idx]:
                        row_dict[headers[col_idx]] = str(value) if value is not None else ""

                # Skip completely empty rows
                if not any(v.strip() for v in row_dict.values() if v):
                    continue

                rows.append({
                    "marker_name": row_dict.get("Marker Name", row_dict.get("marker_name", row_dict.get("Analyte", ""))),
                    "species": row_dict.get("Species", row_dict.get("species", "")),
                    "sample_id": row_dict.get("Sample ID", row_dict.get("sample_id", row_dict.get("Patient ID", ""))),
                    "expected_value": row_dict.get("Expected", row_dict.get("expected_value", row_dict.get("Expected Value", ""))),
                    "measured_value": row_dict.get("Measured", row_dict.get("measured_value", row_dict.get("Measured Value", ""))),
                    "unit": row_dict.get("Unit", row_dict.get("unit", "")),
                    "result_status": row_dict.get("Status", row_dict.get("Result", "")),
                    "sheet_name": sheet_name,
                    "row_index": row_idx,
                    "raw_data": str(row_dict),
                })

        wb.close()
        return rows
    except Exception as exc:
        logger.exception(f"Failed to parse xlsx: {exc}")
        return []


# ─── Store to DB ──────────────────────────────────────────────────────────────

def store_equation(
    parsed: dict,
    table_name: str,
    source_file: str,
    file_hash: str,
    s3_key: str,
    s3_url: str,
) -> dict:
    """Insert or update equation record in tuttiprequn or tuttiprueqn."""
    conn = _get_db()
    try:
        now = datetime.now(timezone.utc).isoformat(timespec="seconds")

        if table_name == "tuttiprueqn":
            # Extract production_date and disc_name from source path
            production_date, disc_name = extract_production_info(source_file)
            sub_panel_type = lookup_sub_panel_type(disc_name)

            conn.execute(
                f"""
                INSERT INTO tuttiprueqn
                    (marker_name, well_number, control_equation, canine_equation,
                     feline_equation, equine_equation, production_date, disc_name,
                     sub_panel_type, source_file, file_hash, s3_key, s3_url,
                     created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(marker_name, source_file) DO UPDATE SET
                    well_number = excluded.well_number,
                    control_equation = excluded.control_equation,
                    canine_equation = excluded.canine_equation,
                    feline_equation = excluded.feline_equation,
                    equine_equation = excluded.equine_equation,
                    production_date = excluded.production_date,
                    disc_name = excluded.disc_name,
                    sub_panel_type = excluded.sub_panel_type,
                    file_hash = excluded.file_hash,
                    s3_key = excluded.s3_key,
                    s3_url = excluded.s3_url,
                    updated_at = excluded.updated_at
                """,
                (
                    parsed["marker_name"],
                    parsed["well_number"],
                    parsed["control_equation"],
                    parsed["canine_equation"],
                    parsed["feline_equation"],
                    parsed["equine_equation"],
                    production_date,
                    disc_name,
                    sub_panel_type,
                    source_file,
                    file_hash,
                    s3_key,
                    s3_url,
                    now,
                    now,
                ),
            )
        else:
            conn.execute(
                f"""
                INSERT INTO {table_name}
                    (marker_name, well_number, control_equation, canine_equation,
                     feline_equation, equine_equation, source_file, file_hash,
                     s3_key, s3_url, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(marker_name, source_file) DO UPDATE SET
                    well_number = excluded.well_number,
                    control_equation = excluded.control_equation,
                    canine_equation = excluded.canine_equation,
                    feline_equation = excluded.feline_equation,
                    equine_equation = excluded.equine_equation,
                    file_hash = excluded.file_hash,
                    s3_key = excluded.s3_key,
                    s3_url = excluded.s3_url,
                    updated_at = excluded.updated_at
                """,
                (
                    parsed["marker_name"],
                    parsed["well_number"],
                    parsed["control_equation"],
                    parsed["canine_equation"],
                    parsed["feline_equation"],
                    parsed["equine_equation"],
                    source_file,
                    file_hash,
                    s3_key,
                    s3_url,
                    now,
                    now,
                ),
            )

        conn.commit()
        result = {"ok": True, "rows_inserted": 1, "marker_name": parsed["marker_name"]}
        if table_name == "tuttiprueqn":
            result["production_date"] = production_date
            result["disc_name"] = disc_name
            result["sub_panel_type"] = sub_panel_type
        return result
    except Exception as exc:
        logger.exception(f"DB insert failed: {exc}")
        return {"ok": False, "error": str(exc)}
    finally:
        conn.close()


def store_real_assign_rows(
    rows: list[dict],
    source_file: str,
    file_hash: str,
    s3_key: str,
    s3_url: str,
) -> dict:
    """Insert rows into tuttirealassign table."""
    if not rows:
        return {"ok": True, "rows_inserted": 0, "message": "no data rows found"}

    conn = _get_db()
    try:
        now = datetime.now(timezone.utc).isoformat(timespec="seconds")
        # Delete old records from same source_file to allow re-import
        conn.execute("DELETE FROM tuttirealassign WHERE source_file = ?", (source_file,))

        inserted = 0
        for row in rows:
            conn.execute(
                """
                INSERT INTO tuttirealassign
                    (marker_name, species, sample_id, expected_value, measured_value,
                     unit, result_status, sheet_name, row_index,
                     source_file, file_hash, s3_key, s3_url, raw_data,
                     created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    row["marker_name"],
                    row["species"],
                    row["sample_id"],
                    row["expected_value"],
                    row["measured_value"],
                    row["unit"],
                    row["result_status"],
                    row["sheet_name"],
                    row["row_index"],
                    source_file,
                    file_hash,
                    s3_key,
                    s3_url,
                    row.get("raw_data", ""),
                    now,
                    now,
                ),
            )
            inserted += 1

        conn.commit()
        return {"ok": True, "rows_inserted": inserted}
    except Exception as exc:
        logger.exception(f"DB insert failed: {exc}")
        return {"ok": False, "error": str(exc)}
    finally:
        conn.close()


def parse_calvalue_xlsx(content: bytes, file_name: str) -> list[dict]:
    """
    Parse Tutti-計算值.xlsx file.
    Returns list of row dicts for calvalue table.
    """
    try:
        import openpyxl
        import io as _io

        wb = openpyxl.load_workbook(_io.BytesIO(content), read_only=True, data_only=True)
        rows = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            all_rows = list(ws.iter_rows(values_only=True))
            if len(all_rows) < 2:
                continue

            headers = [str(h or "").strip() for h in all_rows[0]]

            for row_idx, row_values in enumerate(all_rows[1:], start=2):
                row_dict = {}
                for col_idx, value in enumerate(row_values):
                    if col_idx < len(headers) and headers[col_idx]:
                        row_dict[headers[col_idx]] = str(value) if value is not None else ""

                # Skip completely empty rows
                if not any(v.strip() for v in row_dict.values() if v):
                    continue

                # Try to identify marker_name from common column names
                marker_name = (
                    row_dict.get("Marker Name", "")
                    or row_dict.get("marker_name", "")
                    or row_dict.get("Analyte", "")
                    or row_dict.get("Item", "")
                    or row_dict.get("項目", "")
                )

                rows.append({
                    "marker_name": marker_name,
                    "parameter": row_dict.get("Parameter", row_dict.get("parameter", "")),
                    "value": row_dict.get("Value", row_dict.get("value", row_dict.get("計算值", ""))),
                    "unit": row_dict.get("Unit", row_dict.get("unit", row_dict.get("單位", ""))),
                    "sheet_name": sheet_name,
                    "row_index": row_idx,
                    "raw_data": str(row_dict),
                })

        wb.close()
        return rows
    except Exception as exc:
        logger.exception(f"Failed to parse calvalue xlsx: {exc}")
        return []


def store_calvalue_rows(
    rows: list[dict],
    source_file: str,
    file_hash: str,
    s3_key: str,
    s3_url: str,
) -> dict:
    """Insert rows into calvalue table."""
    if not rows:
        return {"ok": True, "rows_inserted": 0, "message": "no data rows found"}

    conn = _get_db()
    try:
        now = datetime.now(timezone.utc).isoformat(timespec="seconds")
        # Delete old records from same source_file to allow re-import
        conn.execute("DELETE FROM calvalue WHERE source_file = ?", (source_file,))

        inserted = 0
        for row in rows:
            conn.execute(
                """
                INSERT INTO calvalue
                    (marker_name, parameter, value, unit, sheet_name, row_index,
                     source_file, file_hash, s3_key, s3_url, raw_data,
                     created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    row["marker_name"],
                    row["parameter"],
                    row["value"],
                    row["unit"],
                    row["sheet_name"],
                    row["row_index"],
                    source_file,
                    file_hash,
                    s3_key,
                    s3_url,
                    row.get("raw_data", ""),
                    now,
                    now,
                ),
            )
            inserted += 1

        conn.commit()
        return {"ok": True, "rows_inserted": inserted}
    except Exception as exc:
        logger.exception(f"DB insert failed: {exc}")
        return {"ok": False, "error": str(exc)}
    finally:
        conn.close()


# ─── High-level API functions ─────────────────────────────────────────────────

def handle_s3_upload(
    file_content: bytes,
    file_name: str,
    s3_bucket: str | None = None,
    s3_key: str | None = None,
    source_path: str = "",
) -> dict:
    """Handle the /api/tutti/upload-to-s3 request."""
    bucket = s3_bucket or S3_BUCKET
    key = s3_key or f"Tutti/Pur_baseline/{file_name}"

    result = upload_file_to_s3(file_content, bucket, key)
    if result["ok"]:
        result["file_name"] = file_name
        result["source_path"] = source_path
    return result


def handle_analyze_content(
    file_content: bytes,
    file_name: str,
    file_hash: str = "",
    s3_key: str = "",
    s3_url: str = "",
    source_path: str = "",
    db_name: str = "",
    db_table: str = "",
    target_name: str = "",
) -> dict:
    """
    Handle the /api/tutti/analyze-content request.
    Routes to appropriate parser based on target/table.
    """
    # Determine which table we're writing to
    table = db_table or "tuttiprequn"

    if table == "tuttirealassign":
        # XLSX parsing
        rows = parse_real_assign_xlsx(file_content, file_name)
        return store_real_assign_rows(
            rows=rows,
            source_file=source_path or file_name,
            file_hash=file_hash,
            s3_key=s3_key,
            s3_url=s3_url,
        )
    elif table == "calvalue":
        # Tutti-計算值.xlsx parsing
        rows = parse_calvalue_xlsx(file_content, file_name)
        return store_calvalue_rows(
            rows=rows,
            source_file=source_path or file_name,
            file_hash=file_hash,
            s3_key=s3_key,
            s3_url=s3_url,
        )
    else:
        # CSV equation parsing (tuttiprequn or tuttiprueqn)
        parsed = parse_equation_csv(file_content)
        if parsed is None:
            return {
                "ok": False,
                "error": f"Failed to parse equation CSV: {file_name}",
            }
        result = store_equation(
            parsed=parsed,
            table_name=table,
            source_file=source_path or file_name,
            file_hash=file_hash,
            s3_key=s3_key,
            s3_url=s3_url,
        )
        if result["ok"]:
            result["message"] = (
                f"Stored marker '{parsed['marker_name']}' (Well {parsed['well_number']}) "
                f"in {table}"
            )
        return result
