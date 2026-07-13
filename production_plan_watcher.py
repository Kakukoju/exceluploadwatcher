# -*- coding: utf-8 -*-
"""Watch a Production Plan network folder and incrementally sync to EC2/RDS.

Windows setup:
    py -m pip install openpyxl requests
    py production_plan_watcher.py
"""

from __future__ import annotations

import argparse
import hashlib
import json
import logging
import os
import re
import shutil
import tempfile
import time
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl
import requests

# LINE 通知
try:
    from line_notify import send_line_message
except ImportError:
    def send_line_message(msg, source="default"):
        return False


DEFAULT_WATCH_DIR = "/mnt/mbbu_fab/MB_PD/生管自動化/滴定/production_plan"
DEFAULT_API_URL = (
    "https://52-192-28-39.sslip.io/api/upload-production-plan-incremental"
)
DEFAULT_API_KEY = "beadsops-upload-key"
SHEET_NAME = "P_plan Reagent"
HEADER_ROW = 2
FILE_PATTERN = re.compile(r"^Production plan-(\d{8})\.xlsm$", re.IGNORECASE)
ISO_DATE_PATTERN = re.compile(r"^\d{4}-\d{2}-\d{2}$")


def default_state_path() -> Path:
    return Path(__file__).resolve().parent / "production_plan_watcher_state.json"


def normalize_date_header(value: Any) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value or "").strip()
    match = re.match(r"^(\d{4}-\d{2}-\d{2})(?:\s|$)", text)
    return match.group(1) if match else None


def normalize_cell(value: Any) -> Any:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def hash_value(value: Any) -> str:
    encoded = json.dumps(
        value, ensure_ascii=False, sort_keys=True, separators=(",", ":")
    ).encode("utf-8")
    return hashlib.sha256(encoded).hexdigest()


def read_sheet_snapshot(workbook_path: Path) -> dict[str, Any]:
    workbook = openpyxl.load_workbook(
        workbook_path, read_only=True, data_only=True
    )
    try:
        if SHEET_NAME not in workbook.sheetnames:
            raise ValueError(f"找不到工作表: {SHEET_NAME}")
        sheet = workbook[SHEET_NAME]
        row_iterator = sheet.iter_rows(min_row=HEADER_ROW, values_only=True)
        header_values = next(row_iterator)
        panel_column = next(
            (
                index
                for index, value in enumerate(header_values)
                if re.search(r"panel.*no|panel_no", str(value or ""), re.IGNORECASE)
            ),
            None,
        )
        if panel_column is None:
            raise ValueError("找不到 Panel_NO 欄")
        date_columns = {
            index: normalized
            for index, value in enumerate(header_values)
            if (normalized := normalize_date_header(value))
        }
        if not date_columns:
            raise ValueError("找不到日期欄")

        values_by_date: dict[str, list[list[Any]]] = {
            normalized: [] for normalized in date_columns.values()
        }
        panel_numbers: list[str] = []
        for row_values in row_iterator:
            raw_panel = row_values[panel_column]
            if raw_panel is None or str(raw_panel).strip() == "":
                continue
            panel_no = str(normalize_cell(raw_panel)).strip()
            if re.search(r"panel.*no", panel_no, re.IGNORECASE):
                continue
            panel_numbers.append(panel_no)
            for column, normalized in date_columns.items():
                values_by_date[normalized].append(
                    [panel_no, normalize_cell(row_values[column])]
                )

        return {
            "sheet": SHEET_NAME,
            "panel_hash": hash_value(panel_numbers),
            "date_hashes": {
                normalized: hash_value(values)
                for normalized, values in values_by_date.items()
            },
        }
    finally:
        workbook.close()


def earliest_changed_date(
    previous: dict[str, Any] | None,
    current: dict[str, Any],
    initial_since: str,
) -> str | None:
    if not previous:
        available = sorted(
            day
            for day in current["date_hashes"]
            if ISO_DATE_PATTERN.match(day) and day >= initial_since
        )
        return available[0] if available else None

    previous_dates = previous.get("date_hashes", {})
    current_dates = current.get("date_hashes", {})
    changed = sorted(
        day
        for day in set(previous_dates) | set(current_dates)
        if previous_dates.get(day) != current_dates.get(day)
    )
    if changed:
        return changed[0]
    if previous.get("panel_hash") != current.get("panel_hash"):
        return initial_since
    return None


def workbook_filename_date(filename: str) -> str | None:
    match = FILE_PATTERN.match(filename)
    return match.group(1) if match else None


def find_latest_workbook(
    watch_dir: Path, monitored_filename: str | None = None
) -> Path | None:
    candidates: list[tuple[str, Path]] = []
    for path in watch_dir.glob("Production plan-*.xlsm"):
        filename_date = workbook_filename_date(path.name)
        if filename_date:
            candidates.append((filename_date, path))
    if not candidates:
        return None

    latest_date, latest_path = max(candidates, key=lambda item: item[0])
    if not monitored_filename:
        return latest_path

    monitored_date = workbook_filename_date(monitored_filename)
    if monitored_date is None:
        return latest_path
    if latest_date > monitored_date:
        return latest_path

    monitored_path = watch_dir / monitored_filename
    return monitored_path if monitored_path.exists() else None


def file_signature(path: Path) -> dict[str, Any]:
    stat = path.stat()
    return {
        "path": str(path),
        "size": stat.st_size,
        "mtime_ns": stat.st_mtime_ns,
    }


def wait_for_stable_file(
    path: Path, stable_seconds: float, timeout_seconds: float
) -> None:
    deadline = time.monotonic() + timeout_seconds
    previous: tuple[int, int] | None = None
    stable_since: float | None = None
    while time.monotonic() < deadline:
        stat = path.stat()
        current = (stat.st_size, stat.st_mtime_ns)
        if current == previous:
            stable_since = stable_since or time.monotonic()
            if time.monotonic() - stable_since >= stable_seconds:
                return
        else:
            previous = current
            stable_since = None
        time.sleep(1)
    raise TimeoutError(f"檔案在 {timeout_seconds:g} 秒內仍未穩定: {path}")


def make_local_copy(source: Path) -> Path:
    temp_dir = Path(tempfile.mkdtemp(prefix="production-plan-"))
    target = temp_dir / source.name
    shutil.copy2(source, target)
    return target


def upload_workbook(
    workbook_path: Path,
    source_path: Path,
    since_date: str,
    api_url: str,
    api_key: str,
    timeout_seconds: float,
) -> dict[str, Any]:
    with workbook_path.open("rb") as workbook:
        response = requests.post(
            api_url,
            headers={"X-Api-Key": api_key},
            data={"since_date": since_date, "source_path": str(source_path)},
            files={
                "file": (
                    source_path.name,
                    workbook,
                    "application/vnd.ms-excel.sheet.macroEnabled.12",
                )
            },
            timeout=timeout_seconds,
        )
    response.raise_for_status()
    result = response.json()
    if not result.get("ok"):
        raise RuntimeError(result.get("error") or "EC2 回傳同步失敗")
    return result


def load_state(path: Path) -> dict[str, Any]:
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except FileNotFoundError:
        return {}
    except (OSError, json.JSONDecodeError) as exc:
        logging.warning("狀態檔無法讀取，將重建: %s", exc)
        return {}


def save_state(path: Path, state: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_suffix(path.suffix + ".tmp")
    temporary.write_text(
        json.dumps(state, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    temporary.replace(path)


def sync_once(args: argparse.Namespace, state: dict[str, Any]) -> bool:
    monitored_filename = state.get("monitored_filename")
    if not monitored_filename and state.get("source_path"):
        monitored_filename = Path(state["source_path"]).name
    source = find_latest_workbook(
        Path(args.watch_dir), monitored_filename
    )
    if source is None:
        if monitored_filename:
            logging.warning(
                "%s 已刪除；等待日期更新的 Production Plan，不退回舊檔",
                monitored_filename,
            )
        else:
            logging.warning("找不到 Production plan-YYYYMMDD.xlsm")
        return False

    current_signature = file_signature(source)
    if state.get("file_signature") == current_signature:
        logging.debug("%s 檔案未變動，略過 Excel 讀取", source.name)
        return False

    wait_for_stable_file(source, args.stable_seconds, args.stable_timeout)
    current_signature = file_signature(source)
    local_copy = make_local_copy(source)
    try:
        snapshot = read_sheet_snapshot(local_copy)
        since_date = earliest_changed_date(
            state.get("snapshot"), snapshot, args.initial_since
        )
        if since_date is None:
            logging.info("%s 的 %s 沒有資料變動", source.name, SHEET_NAME)
            state.update(
                {
                    "source_path": str(source),
                    "monitored_filename": source.name,
                    "file_signature": current_signature,
                    "snapshot": snapshot,
                    "checked_at": datetime.now().isoformat(timespec="seconds"),
                }
            )
            save_state(Path(args.state_file), state)
            return False

        result = upload_workbook(
            local_copy,
            source,
            since_date,
            args.api_url,
            args.api_key,
            args.upload_timeout,
        )
        state.update(
            {
                "source_path": str(source),
                "monitored_filename": source.name,
                "file_signature": current_signature,
                "snapshot": snapshot,
                "since_date": since_date,
                "synced_at": datetime.now().isoformat(timespec="seconds"),
                "server_result": result,
            }
        )
        save_state(Path(args.state_file), state)
        logging.info(
            "同步完成: %s，從 %s 起，%s 列 / %s 日期欄",
            source.name,
            since_date,
            result.get("rows"),
            result.get("date_columns"),
        )
        return True
    finally:
        shutil.rmtree(local_copy.parent, ignore_errors=True)


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="監控 Production Plan UNC 資料夾並增量同步到 EC2/RDS"
    )
    parser.add_argument(
        "--watch-dir",
        default=os.environ.get("PRODUCTION_PLAN_WATCH_DIR", DEFAULT_WATCH_DIR),
    )
    parser.add_argument(
        "--api-url",
        default=os.environ.get("PRODUCTION_PLAN_API_URL", DEFAULT_API_URL),
    )
    parser.add_argument(
        "--api-key",
        default=os.environ.get("PRODUCTION_PLAN_API_KEY", DEFAULT_API_KEY),
    )
    parser.add_argument(
        "--state-file",
        default=os.environ.get(
            "PRODUCTION_PLAN_STATE_FILE", str(default_state_path())
        ),
    )
    parser.add_argument(
        "--initial-since",
        default=os.environ.get("PRODUCTION_PLAN_INITIAL_SINCE", date.today().isoformat()),
        help="沒有舊快照時，第一次同步的起始日 (YYYY-MM-DD)",
    )
    parser.add_argument(
        "--interval",
        type=float,
        default=300,
        help="檢查資料夾間隔秒數，預設 300 秒",
    )
    parser.add_argument("--stable-seconds", type=float, default=5)
    parser.add_argument("--stable-timeout", type=float, default=120)
    parser.add_argument("--upload-timeout", type=float, default=180)
    parser.add_argument(
        "--once", action="store_true", help="只檢查一次，供測試或排程器使用"
    )
    return parser


def main() -> int:
    args = build_parser().parse_args()
    if not ISO_DATE_PATTERN.match(args.initial_since):
        raise SystemExit("--initial-since 必須是 YYYY-MM-DD")
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)s %(message)s",
    )
    state = load_state(Path(args.state_file))
    logging.info("監控資料夾: %s", args.watch_dir)
    logging.info("狀態檔: %s", args.state_file)

    # Backoff 設定
    max_failures = 3
    backoff_interval = 3600  # 1 小時
    failure_count = 0
    in_backoff = False

    while True:
        try:
            synced = sync_once(args, state)
            # 成功同步 → 重置
            if synced and in_backoff:
                logging.info("[backoff] 同步恢復正常")
                send_line_message(
                    f"✅ Production Plan 同步恢復正常\n"
                    f"檔案: {state.get('monitored_filename', '?')}",
                    source="production_plan",
                )
            if synced:
                failure_count = 0
                in_backoff = False
        except KeyboardInterrupt:
            logging.info("已停止")
            return 0
        except Exception as exc:
            logging.exception("Production Plan 同步失敗，稍後重試")
            failure_count += 1
            if failure_count >= max_failures and not in_backoff:
                in_backoff = True
                logging.warning(
                    "[backoff] 連續失敗 %d 次，進入 backoff 模式 (每 %d 分鐘重試)",
                    failure_count, backoff_interval // 60,
                )
                send_line_message(
                    f"⚠️ Production Plan 同步連續失敗 {failure_count} 次\n"
                    f"錯誤: {str(exc)[:100]}\n"
                    f"已進入 backoff 模式 (每小時重試)",
                    source="production_plan",
                )
        if args.once:
            return 0
        sleep_time = backoff_interval if in_backoff else args.interval
        time.sleep(sleep_time)


if __name__ == "__main__":
    raise SystemExit(main())
