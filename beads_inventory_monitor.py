"""
BEADS 庫存 Excel 監控程式 (Windows 本地端)
- 每 10 分鐘檢查檔案是否變動 (檔名/時間/大小)
- 有變動才讀取 Excel → 上傳 JSON → push to RDS
- 不開啟檔案鎖，不影響工作者
- 連續失敗 3 次後進入 backoff 模式，並透過 LINE 通知

使用方式:
  pip install openpyxl requests
  python beads_inventory_monitor.py

可用 Windows Task Scheduler 設為開機啟動
"""

import os, re, glob, time as time_mod, json, logging, threading
from datetime import datetime, date, time as dt_time
import openpyxl
import requests

try:
    from watchdog.observers.polling import PollingObserver
    from watchdog.events import FileSystemEventHandler
    HAS_WATCHDOG = True
except ImportError:
    HAS_WATCHDOG = False

# ─── 設定 ───
WATCH_DIR = "/mnt/mbbu_fab/MB_PD/生管自動化/滴定/祐銓"
FILE_PATTERN = "*BEADS庫存*-*NEW.xlsm"
API_URL = "https://52-192-28-39.sslip.io/api/upload-beads-json"
API_KEY = "beadsops-upload-key"
INTERVAL = 600  # 10 分鐘
STATE_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".beads_inv_state.json")

SHEET_PREFIX = "BEADS庫存表("
HEADER_ROW = 5
START_COL = 1
END_COL = 15  # A:O

LOG_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "beads_inventory_monitor.log")
LOG_MAX_DAYS = 30

# ─── Backoff 設定 ───
MAX_CONSECUTIVE_FAILURES = 3       # 連續失敗幾次後進入 backoff
BACKOFF_INTERVAL = 3600            # backoff 期間每小時重試一次
BACKOFF_RETRY_RESET_HOURS = 6      # backoff 後每 6 小時重試一次，成功就恢復正常

# ─── LINE Messaging API 設定 ───
LINE_CHANNEL_ACCESS_TOKEN = "zGxbSyzcXGPTGlJA0H9IebOMUFn0XJqgAkI0zt/6UUhUH0HTCm6sgF8vHX2nAe12b5/H2o7YWGSf4iSs7CVrmIXMvoee66U8i6bHcJQVQVuXa5ObdC4bGmeuJnxm0gwnPkrMfMsftT4wKUDeBkLIogdB04t89/1O/w1cDnyilFU="
LINE_USER_ID = "U27153a213b9284361b380b9eb419d069"

# ─── 全域狀態 ───
_failure_count = 0
_in_backoff = False
_last_notified = None  # 避免重複通知


def _rotate_log():
    """log 檔超過 30 天就刪除重建"""
    if os.path.exists(LOG_FILE):
        age = (datetime.now() - datetime.fromtimestamp(os.path.getctime(LOG_FILE))).days
        if age >= LOG_MAX_DAYS:
            os.remove(LOG_FILE)

_rotate_log()
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    handlers=[
        logging.FileHandler(LOG_FILE, encoding="utf-8"),
        logging.StreamHandler(),
    ],
)
log = logging.getLogger(__name__)


# ─── LINE + Teams 通知 ───
TEAMS_WEBHOOK_URL = (
    "https://default15d82f974f154ead9ab618aa0cd453.88.environment.api.powerplatform.com:443"
    "/powerautomate/automations/direct/cu/29/workflows/8e919cca47ed4cb49cd58e589773976c"
    "/triggers/manual/paths/invoke?api-version=1&sp=%2Ftriggers%2Fmanual%2Frun&sv=1.0"
    "&sig=qkNRsez6F5qH6tdMGB6ZNTCYhXtSAU48h2ggTkjBH0s"
)


def send_line_notification(message):
    """透過 LINE + Teams 發送通知"""
    global _last_notified
    now = datetime.now()

    # 同一小時內不重複通知
    if _last_notified and (now - _last_notified).total_seconds() < 3600:
        return

    # LINE
    try:
        resp = requests.post(
            "https://api.line.me/v2/bot/message/push",
            json={"to": LINE_USER_ID, "messages": [{"type": "text", "text": message}]},
            headers={"Content-Type": "application/json", "Authorization": f"Bearer {LINE_CHANNEL_ACCESS_TOKEN}"},
            timeout=10,
        )
        if resp.status_code == 200:
            log.info("[LINE] 通知已發送")
        else:
            log.warning(f"[LINE] 通知失敗: {resp.status_code}")
    except Exception as e:
        log.warning(f"[LINE] 通知發送異常: {e}")

    # Teams (Adaptive Card 格式)
    try:
        teams_payload = {
            "type": "message",
            "attachments": [
                {
                    "contentType": "application/vnd.microsoft.card.adaptive",
                    "content": {
                        "$schema": "http://adaptivecards.io/schemas/adaptive-card.json",
                        "type": "AdaptiveCard",
                        "version": "1.2",
                        "body": [
                            {"type": "TextBlock", "text": message, "wrap": True},
                        ],
                    },
                }
            ],
        }
        resp = requests.post(
            TEAMS_WEBHOOK_URL,
            json=teams_payload,
            headers={"Content-Type": "application/json"},
            timeout=10,
        )
        if resp.status_code in (200, 202):
            log.info("[Teams] 通知已發送")
        else:
            log.warning(f"[Teams] 通知失敗: {resp.status_code}")
    except Exception as e:
        log.warning(f"[Teams] 通知發送異常: {e}")

    _last_notified = now


def find_target_file():
    """找符合 *BEADS庫存YYYYMMDD-YYYYMMDD* 的檔案，多個則取後面日期最新的"""
    pattern = os.path.join(WATCH_DIR, FILE_PATTERN)
    files = glob.glob(pattern)
    # 排除 Excel 鎖定暫存檔 (~$ 開頭)
    files = [f for f in files if not os.path.basename(f).startswith("~$")]
    if not files:
        return None

    # 提取日期對 YYYYMMDD-YYYYMMDD
    date_re = re.compile(r"(\d{8})--?(\d{8})")
    candidates = []
    for f in files:
        m = date_re.search(os.path.basename(f))
        if m:
            candidates.append((m.group(2), f))  # 取後面日期排序

    if not candidates:
        return files[0]  # fallback

    candidates.sort(key=lambda x: x[0], reverse=True)
    return candidates[0][1]


def get_file_state(filepath):
    """取得檔名+修改時間+大小"""
    stat = os.stat(filepath)
    return {
        "path": filepath,
        "name": os.path.basename(filepath),
        "mtime": stat.st_mtime,
        "size": stat.st_size,
    }


def load_prev_state():
    if os.path.exists(STATE_FILE):
        with open(STATE_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return None


def save_state(state):
    with open(STATE_FILE, "w", encoding="utf-8") as f:
        json.dump(state, f, ensure_ascii=False)


def find_sheet(wb):
    """找 BEADS庫存表(YYYYMM~ 的 sheet"""
    for name in wb.sheetnames:
        if name.startswith(SHEET_PREFIX):
            return name
    return None


def read_excel_data(filepath):
    """讀取 Excel A:O, header=row5, start=row6, 過濾全空列"""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    sheet_name = find_sheet(wb)
    if not sheet_name:
        wb.close()
        raise ValueError(f"找不到 sheet 開頭為 '{SHEET_PREFIX}'")

    ws = wb[sheet_name]
    # 讀 header
    headers = []
    for cell in list(ws.iter_rows(min_row=HEADER_ROW, max_row=HEADER_ROW,
                                   min_col=START_COL, max_col=END_COL))[0]:
        headers.append(str(cell.value).strip() if cell.value else f"col{cell.column}")

    # 預期的正確欄名 (A~O)
    EXPECTED_HEADERS = [
        "限制", "PN", "BEADS別", "保存期限 (天)", "Unrestricted",
        "Batch", "生產日期", "併批", "備註", "效期",
        "BEADS工單", "工單數", "入庫", "累計領用", "可使用庫存",
    ]

    # 檢查欄名是否被改動
    if headers != EXPECTED_HEADERS:
        changed = [
            f"  欄{i+1}: '{headers[i]}' → 應為 '{EXPECTED_HEADERS[i]}'"
            for i in range(min(len(headers), len(EXPECTED_HEADERS)))
            if i < len(headers) and headers[i] != EXPECTED_HEADERS[i]
        ]
        if changed:
            msg = (
                f"⚠️ BEADS庫存 Excel 欄名被改動！\n"
                f"檔案: {os.path.basename(filepath)}\n"
                + "\n".join(changed[:5])
                + "\n資料仍以正確欄名上傳"
            )
            log.warning(f"[欄名檢查] 欄名不一致: {changed}")
            send_line_notification(msg)

    # 強制使用正確欄名上傳
    headers = EXPECTED_HEADERS

    # 讀資料
    rows = []
    for row in ws.iter_rows(min_row=HEADER_ROW + 1, min_col=START_COL, max_col=END_COL):
        values = [cell.value for cell in row]
        # 跳過全空列 (所有欄位為 None 或空字串)
        if all(v is None or (isinstance(v, str) and v.strip() == "") for v in values):
            continue
        record = {}
        for i, v in enumerate(values):
            h = headers[i]
            if v is None or (isinstance(v, str) and v.strip() == ""):
                record[h] = None
            elif isinstance(v, datetime):
                record[h] = v.strftime("%Y-%m-%d")
            elif isinstance(v, date):
                record[h] = v.strftime("%Y-%m-%d")
            elif isinstance(v, dt_time):
                record[h] = str(v)
            else:
                record[h] = v
        rows.append(record)

    wb.close()
    log.info(f"讀取 sheet '{sheet_name}': {len(rows)} 筆")
    return rows


def upload_json(data):
    """上傳 JSON 到 API → push to RDS schedule.beads_Inventory"""
    resp = requests.post(
        API_URL,
        json=data,
        headers={
            "Content-Type": "application/json; charset=utf-8",
            "X-Api-Key": API_KEY,
        },
        timeout=60,
    )
    resp.raise_for_status()
    result = resp.json()
    if not result.get("ok"):
        raise RuntimeError(f"API error: {result.get('error')}")
    return result


# ─── 核心: 檢查並上傳 ───
def check_and_upload():
    """檢查檔案變動，有變動就上傳。含 backoff 機制。"""
    global _failure_count, _in_backoff

    try:
        filepath = find_target_file()
        if not filepath:
            log.warning("找不到符合條件的檔案")
            return

        current = get_file_state(filepath)
        prev = load_prev_state()

        changed = (
            prev is None
            or prev.get("name") != current["name"]
            or prev.get("mtime") != current["mtime"]
            or prev.get("size") != current["size"]
        )

        if not changed:
            return

        log.info(f"偵測到變動: {current['name']} (size={current['size']}, mtime={current['mtime']})")
        data = read_excel_data(filepath)
        result = upload_json(data)
        log.info(f"上傳成功: {result.get('rows')} 筆 → schedule.\"beads_Inventory\"")
        save_state(current)

        # 上傳成功 → 重置失敗計數，退出 backoff
        if _in_backoff:
            log.info("[backoff] 上傳恢復正常，退出 backoff 模式")
            send_line_notification(
                f"✅ BEADS庫存上傳恢復正常\n"
                f"檔案: {current['name']}\n"
                f"筆數: {result.get('rows')}"
            )
        _failure_count = 0
        _in_backoff = False

    except PermissionError:
        log.warning("檔案被鎖定中，下次再試")
    except Exception as e:
        log.error(f"錯誤: {e}", exc_info=True)
        _failure_count += 1

        if _failure_count >= MAX_CONSECUTIVE_FAILURES and not _in_backoff:
            _in_backoff = True
            log.warning(
                f"[backoff] 連續失敗 {_failure_count} 次，進入 backoff 模式 "
                f"(每 {BACKOFF_INTERVAL // 60} 分鐘重試一次)"
            )
            send_line_notification(
                f"⚠️ BEADS庫存上傳連續失敗 {_failure_count} 次\n"
                f"錯誤: {str(e)[:100]}\n"
                f"已進入 backoff 模式 (每小時重試)\n"
                f"請檢查 API 狀態"
            )


# ─── Watchdog handler ───
class BeadsFileHandler(FileSystemEventHandler):
    """檔案事件觸發時延遲 5 秒後執行 check（避免存檔過程中讀取）"""
    def __init__(self):
        self._timer = None

    def _debounce(self):
        if self._timer:
            self._timer.cancel()
        self._timer = threading.Timer(5.0, check_and_upload)
        self._timer.start()

    def on_modified(self, event):
        if not event.is_directory and "BEADS庫存" in os.path.basename(event.src_path):
            log.info(f"[watchdog] 偵測到修改: {os.path.basename(event.src_path)}")
            self._debounce()

    def on_created(self, event):
        if not event.is_directory and "BEADS庫存" in os.path.basename(event.src_path):
            log.info(f"[watchdog] 偵測到新檔: {os.path.basename(event.src_path)}")
            self._debounce()


def main():
    log.info("BEADS 庫存監控啟動")
    log.info(f"監控目錄: {WATCH_DIR}")
    log.info(f"檢查間隔: {INTERVAL}s")
    log.info(f"Backoff: 連續 {MAX_CONSECUTIVE_FAILURES} 次失敗後，改為每 {BACKOFF_INTERVAL // 60} 分鐘重試")
    log.info(f"Watchdog: {'啟用 (PollingObserver)' if HAS_WATCHDOG else '未安裝，僅用 polling'}")

    # 啟動 watchdog (網路磁碟用 PollingObserver 較穩定)
    if HAS_WATCHDOG and os.path.isdir(WATCH_DIR):
        observer = PollingObserver(timeout=30)
        observer.schedule(BeadsFileHandler(), WATCH_DIR, recursive=False)
        observer.daemon = True
        observer.start()
        log.info("[watchdog] observer 已啟動")

    # polling: 正常每 10 分鐘，backoff 時每小時
    while True:
        check_and_upload()
        sleep_time = BACKOFF_INTERVAL if _in_backoff else INTERVAL
        time_mod.sleep(sleep_time)


if __name__ == "__main__":
    main()
